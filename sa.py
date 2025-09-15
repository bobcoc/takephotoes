import tkinter as tk
from tkinter import Label, Button, Checkbutton, IntVar, Frame, OptionMenu, StringVar
from PIL import Image, ImageTk
import cv2
import openpyxl
import subprocess
import os
import signal
import threading
import atexit

def load_students_info(excel_path, sheet_name=None):
    """从 Excel 文件读取学生信息"""
    workbook = openpyxl.load_workbook(excel_path)
    if sheet_name:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
    students_info = []
    for row in sheet.iter_rows(min_row=2):
        exam_id = row[0].value
        name = row[1].value
        if exam_id and name:
            students_info.append((exam_id, name))
    return students_info

def get_sheet_names(excel_path):
    """获取 Excel 文件中所有 sheet 的名称"""
    workbook = openpyxl.load_workbook(excel_path)
    return workbook.sheetnames

class CameraApp:
    def __init__(self, master, excel_path):
        self.master = master
        self.excel_path = excel_path
        self.sheet_names = get_sheet_names(excel_path)
        self.sheet_var = StringVar(master)
        self.sheet_var.set(self.sheet_names[0])  # 默认选择第一个 sheet
        self.students_info = load_students_info(excel_path, self.sheet_var.get())
        self.current_student_index = 0
        self.ffmpeg_process = None
        self.mode_var = IntVar(value=1)  # 默认选中录像模式
        self.master.title("学生录像系统")
        self.master.geometry("1000x750")  # 增加窗口大小

        self.vid = cv2.VideoCapture(0)
        self.vid.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
        self.vid.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)
        self.canvas = tk.Canvas(master, width=960, height=540)  # 增加画布大小
        self.canvas.pack(pady=10)
        
        self.label = Label(master, text="", font=("Arial", 12))
        self.label.pack(pady=5)

        # 创建一个框架来容纳按钮，并使用水平布局
        button_frame = Frame(master)
        button_frame.pack(pady=10)

        self.btn_snapshot = Button(button_frame, text="开始/停止录像", command=self.toggle_recording, width=15, height=2)
        self.btn_snapshot.pack(side=tk.LEFT, padx=5)

        self.btn_previous = Button(button_frame, text="上一个学生", command=self.previous_student, width=15, height=2)
        self.btn_previous.pack(side=tk.LEFT, padx=5)

        self.btn_next = Button(button_frame, text="下一个学生", command=self.next_student, width=15, height=2)
        self.btn_next.pack(side=tk.LEFT, padx=5)

        self.chk_mode = Checkbutton(button_frame, text="录像模式", variable=self.mode_var, command=self.toggle_mode, width=10, height=2)
        self.chk_mode.pack(side=tk.LEFT, padx=5)

        # 添加 sheet 选择下拉框
        sheet_frame = Frame(master)
        sheet_frame.pack(pady=10)
        
        Label(sheet_frame, text="选择 Sheet：").pack(side=tk.LEFT)
        self.sheet_menu = OptionMenu(sheet_frame, self.sheet_var, *self.sheet_names, command=self.change_sheet)
        self.sheet_menu.pack(side=tk.LEFT)

        self.update_student_info()
        self.update()

        # 注册清理函数
        atexit.register(self.cleanup)

    def change_sheet(self, *args):
        """更改选中的 sheet 并重新加载学生信息"""
        self.students_info = load_students_info(self.excel_path, self.sheet_var.get())
        self.current_student_index = 0
        self.update_student_info()

    # ... [其他方法保持不变] ...

    def toggle_mode(self):
        """根据复选框切换模式"""
        if self.ffmpeg_process:
            self.stop_recording()

    def toggle_recording(self):
        """根据模式开始或停止录像"""
        if self.mode_var.get() == 1:  # 录像模式
            if self.ffmpeg_process:
                self.stop_recording()
                self.next_student()
            else:
                self.start_recording()
        else:  # 拍照模式
            self.take_snapshot()
            self.next_student()

    def start_recording(self):
        """开始录像"""
        exam_id, name = self.students_info[self.current_student_index]
        video_name = f"{exam_id}_{name}.mp4"
        
        command = [
            "ffmpeg",
            "-f", "avfoundation",
            "-framerate", "30",
            "-video_size", "1280x720",
            "-i", "0:0",  # 使用默认视频和音频设备
            "-c:v", "libx264",
            "-preset", "ultrafast",
            "-c:a", "aac",
            "-movflags", "+faststart",
            "-y",
            video_name
        ]
        
        self.ffmpeg_process = subprocess.Popen(command, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        print(f"Recording started for {name} ({exam_id}).")

    def stop_recording(self):
        """停止录像并保存文件"""
        if self.ffmpeg_process:
            self.ffmpeg_process.send_signal(signal.SIGINT)
            try:
                self.ffmpeg_process.wait(timeout=5)
            except subprocess.TimeoutExpired:
                self.ffmpeg_process.kill()
                self.ffmpeg_process.wait()
            self.ffmpeg_process = None
            exam_id, name = self.students_info[self.current_student_index]
            print(f"Recording stopped and saved for {name} ({exam_id}).")

    def take_snapshot(self):
        """拍照功能"""
        ret, frame = self.vid.read()
        if ret:
            exam_id, name = self.students_info[self.current_student_index]
            cv2.imwrite(f"{exam_id}_{name}.png", frame)
            print(f"Photo saved as {exam_id}_{name}.png")

    def next_student(self):
        """切换到下一个学生"""
        if self.current_student_index < len(self.students_info) - 1:
            if self.ffmpeg_process:
                self.stop_recording()
            self.current_student_index += 1
            self.update_student_info()

    def previous_student(self):
        """切换到上一个学生"""
        if self.current_student_index > 0:
            if self.ffmpeg_process:
                self.stop_recording()
            self.current_student_index -= 1
            self.update_student_info()

    def update_student_info(self):
        """更新当前学生信息"""
        if self.students_info:
            exam_id, name = self.students_info[self.current_student_index]
            self.label.config(text=f"当前学生：{name} ({exam_id})")
        else:
            self.label.config(text="没有学生信息")

    def update(self):
        """更新画布上的图像"""
        ret, frame = self.vid.read()
        if ret:
            frame_resized = cv2.resize(frame, (960, 540))  # 调整帧大小
            self.photo = ImageTk.PhotoImage(image=Image.fromarray(cv2.cvtColor(frame_resized, cv2.COLOR_BGR2RGB)))
            self.canvas.create_image(0, 0, image=self.photo, anchor=tk.NW)
        self.master.after(10, self.update)

    def cleanup(self):
        """清理资源"""
        print("Cleaning up resources...")
        if self.ffmpeg_process:
            self.stop_recording()
        if self.vid.isOpened():
            self.vid.release()
        cv2.destroyAllWindows()
        print("Cleanup completed.")

    def on_closing(self):
        """窗口关闭时的处理函数"""
        print("Window is closing. Cleaning up...")
        self.cleanup()
        self.master.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    excel_path = "mt.xlsx"  # 确保这是您的 Excel 文件的正确路径
    app = CameraApp(root, excel_path)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    
    try:
        root.mainloop()
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        print("Program is exiting. Performing final cleanup...")
        app.cleanup()

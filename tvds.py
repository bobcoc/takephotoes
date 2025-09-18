import tkinter as tk
from tkinter import Label, Button, Entry, IntVar, Frame, Checkbutton, ttk
from PIL import Image, ImageTk
import cv2
import openpyxl
import subprocess
import os
import signal
import threading
import queue
import atexit
import traceback
import numpy as np
import time

def load_students_info(excel_path, sheet_index=0):
    print(f"Loading students info from sheet index: {sheet_index}")
    workbook = openpyxl.load_workbook(excel_path, read_only=True)
    sheets = workbook.sheetnames
    if 0 <= sheet_index < len(sheets):
        sheet = workbook[sheets[sheet_index]]
    else:
        print(f"Invalid sheet index: {sheet_index}. Using the first sheet.")
        sheet = workbook.active
    students_info = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        exam_id, name = row[0], row[1]
        if exam_id and name:
            students_info.append((exam_id, name))
    print(f"Loaded {len(students_info)} students")
    workbook.close()
    return students_info

def get_sheet_names(excel_path):
    """获取Excel文件中所有sheet的名称"""
    workbook = openpyxl.load_workbook(excel_path, read_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    return sheet_names

def detect_cameras():
    """检测可用的摄像头"""
    available_cameras = []
    
    print("正在检测可用的摄像头...")
    
    # 检测前10个可能的摄像头索引
    for index in range(10):
        cap = cv2.VideoCapture(index)
        if cap.isOpened():
            # 尝试读取一帧来确认摄像头真的可用
            ret, frame = cap.read()
            if ret:
                # 获取摄像头信息
                width = cap.get(cv2.CAP_PROP_FRAME_WIDTH)
                height = cap.get(cv2.CAP_PROP_FRAME_HEIGHT)
                fps = cap.get(cv2.CAP_PROP_FPS)
                
                camera_info = {
                    'index': index,
                    'name': f"摄像头 {index} ({int(width)}x{int(height)})",
                    'width': int(width),
                    'height': int(height),
                    'fps': fps
                }
                available_cameras.append(camera_info)
                print(f"✅ 摄像头 {index}: {int(width)}x{int(height)} @ {fps:.1f}fps")
            cap.release()
        else:
            # 如果连续3个索引都没有摄像头，就停止检测
            if index > 2 and len(available_cameras) == 0:
                break
    
    if not available_cameras:
        print("❌ 未检测到任何可用的摄像头")
        # 如果没有检测到摄像头，添加默认选项
        available_cameras.append({
            'index': 0,
            'name': "默认摄像头 (索引 0)",
            'width': 640,
            'height': 480,
            'fps': 30.0
        })
    else:
        print(f"\n共检测到 {len(available_cameras)} 个可用摄像头")
    
    return available_cameras

class CameraApp:
    def __init__(self, master, excel_path):
        self.master = master
        self.excel_path = excel_path
        self.students_info = []
        self.current_student_index = 0
        self.ffmpeg_process = None
        self.sheet_names = []
        self.current_sheet_index = 0
        self.available_cameras = []
        self.current_camera_index = 0
        self.master.title("学生录像系统")
        self.master.geometry("1200x900")

        # 检测可用摄像头
        self.available_cameras = detect_cameras()
        
        # 初始化摄像头
        self.init_camera()
        
        # 获取摄像头的实际帧率
        self.camera_fps = self.vid.get(cv2.CAP_PROP_FPS)
        if self.camera_fps <= 0 or self.camera_fps > 60:
            self.camera_fps = 30.0  # 默认帧率
        print(f"摄像头帧率: {self.camera_fps} fps")
        
        # 计算每帧的时间间隔（毫秒）
        self.frame_interval = int(1000 / self.camera_fps)
        print(f"帧间隔: {self.frame_interval} ms")

        # 创建一个主框架来包含所有元素
        self.main_frame = Frame(master)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # 创建画布框架，并使其能够扩展
        self.canvas_frame = Frame(self.main_frame)
        self.canvas_frame.pack(fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(self.canvas_frame)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        
        self.label = Label(self.main_frame, text="加载中...", font=("Arial", 12))
        self.label.pack(pady=5)

        self.recording_status = Label(self.main_frame, text="就绪", font=("Arial", 12), fg="green")
        self.recording_status.pack(pady=5)

        button_frame = Frame(self.main_frame)
        button_frame.pack(pady=10)

        self.btn_recording = Button(button_frame, text="开始录像", command=self.toggle_recording, width=15, height=2)
        self.btn_recording.pack(side=tk.LEFT, padx=5)

        self.btn_snapshot = Button(button_frame, text="拍照", command=self.take_snapshot, width=15, height=2)
        self.btn_snapshot.pack(side=tk.LEFT, padx=5)

        self.btn_previous = Button(button_frame, text="上一个学生", command=self.previous_student, width=15, height=2)
        self.btn_previous.pack(side=tk.LEFT, padx=5)

        self.btn_next = Button(button_frame, text="下一个学生", command=self.next_student, width=15, height=2)
        self.btn_next.pack(side=tk.LEFT, padx=5)

        self.rotate_var = IntVar(value=0)
        self.chk_rotate = Checkbutton(button_frame, text="旋转180度", variable=self.rotate_var, command=self.toggle_rotation, width=10, height=2)
        self.chk_rotate.pack(side=tk.LEFT, padx=5)

        # 班级选择控件直接添加到按钮框架中
        Label(button_frame, text="班级：", font=("Arial", 10)).pack(side=tk.LEFT, padx=(10, 2))
        self.class_var = tk.StringVar()
        self.class_combo = ttk.Combobox(button_frame, textvariable=self.class_var, width=12, state="readonly", font=("Arial", 10))
        self.class_combo.pack(side=tk.LEFT, padx=2)
        self.class_combo.bind("<<ComboboxSelected>>", self.on_class_selected)

        # 摄像头选择控件
        Label(button_frame, text="摄像头：", font=("Arial", 10)).pack(side=tk.LEFT, padx=(10, 2))
        self.camera_var = tk.StringVar()
        self.camera_combo = ttk.Combobox(button_frame, textvariable=self.camera_var, width=20, state="readonly", font=("Arial", 10))
        self.camera_combo.pack(side=tk.LEFT, padx=2)
        self.camera_combo.bind("<<ComboboxSelected>>", self.on_camera_selected)
        
        # 填充摄像头选项
        camera_names = [cam['name'] for cam in self.available_cameras]
        self.camera_combo['values'] = camera_names
        if camera_names:
            self.camera_combo.set(camera_names[0])  # 设置默认选择第一个摄像头


        self.queue = queue.Queue()
        atexit.register(self.cleanup)

        self.master.after(100, self.load_excel_data)
        self.master.after(self.frame_interval, self.update)
        self.master.after(100, self.process_queue)

        # 绑定窗口大小变化事件
        self.master.bind("<Configure>", self.on_resize)

    def on_resize(self, event):
        # 当窗口大小改变时，调整画布大小
        self.update_canvas_size()

    def init_camera(self):
        """初始化摄像头"""
        if self.available_cameras:
            camera_index = self.available_cameras[0]['index']
            print(f"初始化摄像头，索引: {camera_index}")
        else:
            camera_index = 0
            print("使用默认摄像头，索引: 0")
        
        self.vid = cv2.VideoCapture(camera_index)
        self.vid.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
        self.vid.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)
        self.current_camera_index = camera_index

    def switch_camera(self, camera_index):
        """切换摄像头"""
        # 停止当前录像（如果正在录像）
        if hasattr(self, 'is_recording') and self.is_recording:
            self.stop_recording()
        
        # 释放当前摄像头
        if hasattr(self, 'vid') and self.vid.isOpened():
            self.vid.release()
        
        # 初始化新摄像头
        print(f"切换到摄像头，索引: {camera_index}")
        self.vid = cv2.VideoCapture(camera_index)
        self.vid.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
        self.vid.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)
        self.current_camera_index = camera_index
        
        # 更新帧率信息
        self.camera_fps = self.vid.get(cv2.CAP_PROP_FPS)
        if self.camera_fps <= 0 or self.camera_fps > 60:
            self.camera_fps = 30.0
        self.frame_interval = int(1000 / self.camera_fps)
        print(f"新摄像头帧率: {self.camera_fps} fps")

    def on_camera_selected(self, event):
        """当摄像头选择改变时的回调函数"""
        selected_camera = self.camera_var.get()
        if selected_camera:
            # 查找选中的摄像头索引
            for camera in self.available_cameras:
                if camera['name'] == selected_camera:
                    camera_index = camera['index']
                    if camera_index != self.current_camera_index:
                        self.switch_camera(camera_index)
                    break

    def update_canvas_size(self):
        # 获取主窗口的当前大小
        window_width = self.master.winfo_width()
        window_height = self.master.winfo_height()

        # 计算画布应该的大小（减去按钮和标签的高度）
        canvas_height = window_height - 150  # 恢复原来的高度，因为班级选择框已合并到按钮行
        canvas_width = window_width - 20  # 留一些边距

        # 设置画布大小
        self.canvas.config(width=canvas_width, height=canvas_height)

    def load_excel_data(self):
        print("Starting to load Excel data")
        threading.Thread(target=self._load_excel_data_thread, daemon=True).start()

    def _load_excel_data_thread(self):
        try:
            # 首先获取所有sheet名称
            sheet_names = get_sheet_names(self.excel_path)
            self.queue.put(("update_sheets", sheet_names))
            
            # 然后加载第一个sheet的学生信息
            students_info = load_students_info(self.excel_path, 0)
            self.queue.put(("update_students", students_info))
        except Exception as e:
            print(f"Error in _load_excel_data_thread: {e}")
            traceback.print_exc()
            self.queue.put(("error", str(e)))
        finally:
            self.queue.put(("done", None))

    def process_queue(self):
        try:
            while True:
                message, data = self.queue.get_nowait()
                if message == "update_sheets":
                    print(f"Updating sheet names: {data}")
                    self.sheet_names = data
                    self.class_combo['values'] = data
                    if data:
                        self.class_combo.set(data[0])  # 设置默认选择第一个班级
                elif message == "update_students":
                    print(f"Updating students: {len(data)} students loaded")
                    self.students_info = data
                    self.update_student_info()
                elif message == "error":
                    print(f"An error occurred: {data}")
                elif message == "done":
                    print("Finished processing queue")
                    break
        except queue.Empty:
            pass
        finally:
            self.master.after(100, self.process_queue)


    def toggle_rotation(self):
        print(f"Rotation toggled: {'开启' if self.rotate_var.get() == 1 else '关闭'}")

    def on_class_selected(self, event):
        """当班级选择改变时的回调函数"""
        selected_class = self.class_var.get()
        if selected_class and selected_class in self.sheet_names:
            sheet_index = self.sheet_names.index(selected_class)
            print(f"Selected class: {selected_class}, sheet index: {sheet_index}")
            self.current_sheet_index = sheet_index
            self.current_student_index = 0  # 重置到第一个学生
            self.label.config(text="正在加载...")
            threading.Thread(target=self._load_class_students, args=(sheet_index,), daemon=True).start()

    def _load_class_students(self, sheet_index):
        """加载指定班级的学生信息"""
        try:
            students_info = load_students_info(self.excel_path, sheet_index)
            self.queue.put(("update_students", students_info))
        except Exception as e:
            print(f"Error in _load_class_students: {e}")
            traceback.print_exc()
            self.queue.put(("error", str(e)))
        finally:
            self.queue.put(("done", None))

    def toggle_recording(self):
        if hasattr(self, 'is_recording') and self.is_recording:
            self.stop_recording()
            self.next_student()
        else:
            self.start_recording()

    def start_recording(self):
        exam_id, name = self.students_info[self.current_student_index]
        video_name = f"{exam_id}_{name}.mp4"
        
        # 创建一个管道来传输视频帧，包含时间戳信息
        self.frame_queue = queue.Queue(maxsize=100)
        
        # 记录录制开始时间，用于精确的时间戳控制
        self.recording_start_time = time.time()
        self.frame_count = 0
        
        # 启动一个新线程来处理视频帧
        self.recording_thread = threading.Thread(target=self.process_frames, args=(video_name,))
        self.recording_thread.start()
        
        self.is_recording = True
        self.recording_status.config(text="正在录像", fg="red")
        self.btn_recording.config(text="结束录像")
        print(f"Recording started for {name} ({exam_id}). Target FPS: {self.camera_fps}")

    def process_frames(self, video_name):
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        # 使用摄像头的实际帧率
        out = cv2.VideoWriter(video_name, fourcc, self.camera_fps, (1920, 1080))
        
        frame_duration = 1.0 / self.camera_fps  # 每帧应该的时间间隔（秒）
        last_write_time = self.recording_start_time
        frames_written = 0
        frames_received = 0
        
        print(f"开始录制，目标帧率: {self.camera_fps} fps，帧间隔: {frame_duration:.3f}s")
        
        while self.is_recording:
            try:
                frame_data = self.frame_queue.get(timeout=1)
                frames_received += 1
                
                # 计算当前应该写入的时间点
                target_time = self.recording_start_time + frames_written * frame_duration
                current_time = time.time()
                
                # 如果当前时间已经超过了目标时间，说明需要写入帧
                if current_time >= target_time:
                    out.write(frame_data)
                    frames_written += 1
                    last_write_time = current_time
                    
                    # 每100帧打印一次统计信息
                    if frames_written % 100 == 0:
                        actual_fps = frames_written / (current_time - self.recording_start_time)
                        print(f"已录制 {frames_written} 帧，实际FPS: {actual_fps:.2f}，队列大小: {self.frame_queue.qsize()}")
                else:
                    # 如果时间还没到，将帧放回队列
                    self.frame_queue.put(frame_data)
                    time.sleep(0.001)  # 短暂休眠避免CPU占用过高
                    
            except queue.Empty:
                continue
        
        # 录制结束统计
        total_time = time.time() - self.recording_start_time
        actual_fps = frames_written / total_time if total_time > 0 else 0
        print(f"录制结束: 总时长 {total_time:.2f}s，写入帧数 {frames_written}，实际FPS {actual_fps:.2f}")
        
        out.release()

    def stop_recording(self):
        if hasattr(self, 'is_recording') and self.is_recording:
            self.is_recording = False
            self.recording_thread.join()
            exam_id, name = self.students_info[self.current_student_index]
            
            # 计算录制统计信息
            if hasattr(self, 'recording_start_time'):
                total_recording_time = time.time() - self.recording_start_time
                captured_fps = self.frame_count / total_recording_time if total_recording_time > 0 else 0
                print(f"Recording stopped for {name} ({exam_id}).")
                print(f"录制统计: 总时长 {total_recording_time:.2f}s，捕获帧数 {self.frame_count}，捕获FPS {captured_fps:.2f}")
            else:
                print(f"Recording stopped and saved for {name} ({exam_id}).")
            
            self.recording_status.config(text="就绪", fg="green")
            self.btn_recording.config(text="开始录像")

    def take_snapshot(self):
        ret, frame = self.vid.read()
        if ret:
            if self.rotate_var.get() == 1:
                frame = cv2.rotate(frame, cv2.ROTATE_180)
            exam_id, name = self.students_info[self.current_student_index]
            photo_name = f"{exam_id}_{name}.png"
            cv2.imwrite(photo_name, frame)
            print(f"Photo saved as {photo_name}")
            # 如果正在录像，显示拍照提示
            if hasattr(self, 'is_recording') and self.is_recording:
                print(f"Photo taken during recording for {name} ({exam_id})")

    def next_student(self):
        if self.current_student_index < len(self.students_info) - 1:
            if hasattr(self, 'is_recording') and self.is_recording:
                self.stop_recording()
            self.current_student_index += 1
            self.update_student_info()

    def previous_student(self):
        if self.current_student_index > 0:
            if hasattr(self, 'is_recording') and self.is_recording:
                self.stop_recording()
            self.current_student_index -= 1
            self.update_student_info()

    def update_student_info(self):
        if self.students_info:
            exam_id, name = self.students_info[self.current_student_index]
            self.label.config(text=f"当前学生：{name} ({exam_id})")
        else:
            self.label.config(text="没有学生信息")

    def update(self):
        ret, frame = self.vid.read()
        if ret:
            if self.rotate_var.get() == 1:
                frame = cv2.rotate(frame, cv2.ROTATE_180)
            
            # 如果正在录像，将帧添加到队列中，包含时间戳
            if hasattr(self, 'is_recording') and self.is_recording:
                if self.frame_queue.qsize() < 100:  # 限制队列大小以防内存溢出
                    # 确保帧的尺寸正确
                    frame_for_recording = cv2.resize(frame, (1920, 1080))
                    self.frame_queue.put(frame_for_recording)
                    self.frame_count += 1
            
            # 获取当前画布大小
            canvas_width = self.canvas.winfo_width()
            canvas_height = self.canvas.winfo_height()

            # 调整frame大小以适应画布，保持宽高比
            frame_height, frame_width = frame.shape[:2]
            aspect_ratio = frame_width / frame_height
            
            if canvas_width / canvas_height > aspect_ratio:
                new_height = canvas_height
                new_width = int(new_height * aspect_ratio)
            else:
                new_width = canvas_width
                new_height = int(new_width / aspect_ratio)

            frame_resized = cv2.resize(frame, (new_width, new_height))
            self.photo = ImageTk.PhotoImage(image=Image.fromarray(cv2.cvtColor(frame_resized, cv2.COLOR_BGR2RGB)))
            
            # 清除之前的图像并创建新的图像
            self.canvas.delete("all")
            self.canvas.create_image(canvas_width//2, canvas_height//2, image=self.photo, anchor=tk.CENTER)

        self.master.after(self.frame_interval, self.update)

    def cleanup(self):
        print("Cleaning up resources...")
        if hasattr(self, 'is_recording') and self.is_recording:
            self.stop_recording()
        if self.vid.isOpened():
            self.vid.release()
        cv2.destroyAllWindows()
        print("Cleanup completed.")

    def on_closing(self):
        print("Window is closing. Cleaning up...")
        self.cleanup()
        self.master.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    excel_path = "mt2025.xlsx"  # 使用转换后的文件
    app = CameraApp(root, excel_path)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    
    try:
        root.mainloop()
    except Exception as e:
        print(f"An error occurred in main loop: {e}")
        traceback.print_exc()
    finally:
        print("Program is exiting. Performing final cleanup...")
        app.cleanup()
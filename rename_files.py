#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
学生文件重命名工具
根据修改后的 mt2025.xlsx 文件重命名现有的 PNG 和 MP4 文件
"""

import os
import glob
import openpyxl
import re
from pathlib import Path

def load_all_students_from_excel(excel_path):
    """从Excel文件的所有sheet中加载学生信息"""
    print(f"正在读取Excel文件: {excel_path}")
    workbook = openpyxl.load_workbook(excel_path, read_only=True)
    all_students = {}  # 使用字典，key为姓名，value为考号
    
    for sheet_name in workbook.sheetnames:
        print(f"正在处理sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2 and row[0] and row[1]:
                exam_id, name = str(row[0]).strip(), str(row[1]).strip()
                if name and exam_id:
                    all_students[name] = exam_id
                    print(f"  添加学生: {name} -> {exam_id}")
    
    workbook.close()
    print(f"总共加载了 {len(all_students)} 个学生信息")
    return all_students

def extract_name_from_filename(filename):
    """从文件名中提取学生姓名"""
    # 移除文件扩展名
    name_part = os.path.splitext(filename)[0]
    
    # 匹配模式：考号_姓名 或 姓名
    # 先尝试匹配 "数字_姓名" 的模式
    match = re.match(r'^\d+_(.+)$', name_part)
    if match:
        return match.group(1)
    
    # 如果没有匹配到，可能文件名就是姓名
    return name_part

def find_files_to_rename(directory, extensions):
    """查找需要重命名的文件"""
    files_info = []
    
    for ext in extensions:
        pattern = os.path.join(directory, f"*.{ext}")
        files = glob.glob(pattern)
        
        for file_path in files:
            filename = os.path.basename(file_path)
            name = extract_name_from_filename(filename)
            
            files_info.append({
                'path': file_path,
                'filename': filename,
                'name': name,
                'extension': ext
            })
    
    return files_info

def rename_files(directory, excel_path, dry_run=True):
    """重命名文件的主函数"""
    print("="*60)
    print("学生文件重命名工具")
    print("="*60)
    
    # 加载学生信息
    students_dict = load_all_students_from_excel(excel_path)
    if not students_dict:
        print("❌ 没有从Excel文件中读取到学生信息")
        return
    
    # 查找需要重命名的文件
    print(f"\n正在扫描目录: {directory}")
    files_to_process = find_files_to_rename(directory, ['png', 'mp4'])
    
    if not files_to_process:
        print("❌ 没有找到需要处理的PNG或MP4文件")
        return
    
    print(f"找到 {len(files_to_process)} 个文件需要处理")
    
    # 处理文件重命名
    renamed_count = 0
    not_found_count = 0
    already_correct_count = 0
    
    print(f"\n{'模式' if dry_run else '执行模式'}: {'预览重命名操作' if dry_run else '实际执行重命名'}")
    print("-" * 60)
    
    for file_info in files_to_process:
        old_path = file_info['path']
        old_filename = file_info['filename']
        student_name = file_info['name']
        extension = file_info['extension']
        
        # 在学生字典中查找对应的考号
        if student_name in students_dict:
            new_exam_id = students_dict[student_name]
            new_filename = f"{new_exam_id}_{student_name}.{extension}"
            new_path = os.path.join(directory, new_filename)
            
            # 检查是否需要重命名
            if old_filename == new_filename:
                print(f"✅ 文件名已正确: {old_filename}")
                already_correct_count += 1
            else:
                print(f"🔄 重命名: {old_filename} -> {new_filename}")
                
                if not dry_run:
                    try:
                        # 检查目标文件是否已存在
                        if os.path.exists(new_path):
                            print(f"⚠️  目标文件已存在，跳过: {new_filename}")
                            continue
                        
                        os.rename(old_path, new_path)
                        print(f"✅ 重命名成功")
                        renamed_count += 1
                    except Exception as e:
                        print(f"❌ 重命名失败: {e}")
                else:
                    renamed_count += 1
        else:
            print(f"❌ 在Excel中未找到学生: {student_name} (文件: {old_filename})")
            not_found_count += 1
    
    # 输出统计信息
    print("\n" + "="*60)
    print("处理结果统计:")
    print(f"  文件名已正确: {already_correct_count}")
    print(f"  {'预计重命名' if dry_run else '实际重命名'}: {renamed_count}")
    print(f"  学生未找到: {not_found_count}")
    print(f"  总计处理: {len(files_to_process)}")
    
    if dry_run and renamed_count > 0:
        print(f"\n💡 这是预览模式。如需实际执行重命名，请将 dry_run 参数设置为 False")

def main():
    """主函数"""
    # 设置路径
    current_dir = os.getcwd()
    excel_path = os.path.join(current_dir, "mt2025.xlsx")
    
    # 检查Excel文件是否存在
    if not os.path.exists(excel_path):
        print(f"❌ Excel文件不存在: {excel_path}")
        return
    
    print(f"工作目录: {current_dir}")
    print(f"Excel文件: {excel_path}")
    
    # 首先进行预览
    print("\n" + "="*60)
    print("第一步: 预览重命名操作")
    print("="*60)
    rename_files(current_dir, excel_path, dry_run=True)
    
    # 询问用户是否继续
    print("\n" + "="*60)
    response = input("是否继续执行实际重命名操作? (y/N): ").strip().lower()
    
    if response in ['y', 'yes', '是']:
        print("\n第二步: 执行实际重命名")
        print("="*60)
        rename_files(current_dir, excel_path, dry_run=False)
        print("\n✅ 重命名操作完成!")
    else:
        print("\n❌ 操作已取消")

if __name__ == "__main__":
    main()
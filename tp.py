import tkinter as tk
from tkinter import Label, Button, Entry, IntVar, Frame, Checkbutton, Scale, messagebox
from PIL import Image, ImageTk
import cv2
import openpyxl
import subprocess
import os
import signal
import threading
import queue
import atexit
import traceback
import time

def load_students_info(excel_path, sheet_index=0):
    print(f"Loading students info from sheet index: {sheet_index}")
    workbook = openpyxl.load_workbook(excel_path, read_only=True)
    sheets = workbook.sheetnames
    if 0 <= sheet_index < len(sheets):
        sheet = workbook[sheets[sheet_index]]
    else:
        print(f"Invalid sheet index: {sheet_index}. Using the first sheet.")
        sheet = workbook.active
    students_info = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        exam_id, name = row[0], row[1]
        if exam_id and name:
            students_info.append((exam_id, name))
    print(f"Loaded {len(students_info)} students")
    workbook.close()
    return students_info

class CameraApp:
    def __init__(self, master, excel_path):
        self.master = master
        self.excel_path = excel_path
        self.students_info = []
        self.current_student_index = 0
        self.ffmpeg_process = None
        self.is_recording = False
        self.mode_var = IntVar(value=1)
        self.master.title("学生录像系统")
        self.master.geometry("1000x800")

        self.vid = cv2.VideoCapture(0)
        self.vid.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
        self.vid.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)

        self.main_frame = Frame(master)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        self.canvas_frame = Frame(self.main_frame)
        self.canvas_frame.pack(fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(self.canvas_frame)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        
        control_frame = Frame(self.main_frame)
        control_frame.pack(fill=tk.X, padx=10, pady=5)

        self.label = Label(control_frame, text="加载中...", font=("Arial", 12))
        self.label.pack(side=tk.LEFT, padx=5)

        self.recording_status = Label(control_frame, text="就绪", font=("Arial", 12), fg="green")
        self.recording_status.pack(side=tk.LEFT, padx=5)

        sheet_frame = Frame(control_frame)
        sheet_frame.pack(side=tk.RIGHT, padx=5)
        
        Label(sheet_frame, text="Sheet索引：").pack(side=tk.LEFT)
        self.sheet_entry = Entry(sheet_frame, width=5)
        self.sheet_entry.insert(0, "0")
        self.sheet_entry.pack(side=tk.LEFT, padx=2)

        self.btn_load_sheet = Button(sheet_frame, text="加载Sheet", command=self.load_sheet, width=10)
        self.btn_load_sheet.pack(side=tk.LEFT, padx=2)

        button_frame = Frame(self.main_frame)
        button_frame.pack(fill=tk.X, padx=10, pady=5)

        self.btn_snapshot = Button(button_frame, text="开始/停止录像", command=self.toggle_recording, width=15)
        self.btn_snapshot.pack(side=tk.LEFT, padx=5)

        self.btn_previous = Button(button_frame, text="上一个学生", command=self.previous_student, width=15)
        self.btn_previous.pack(side=tk.LEFT, padx=5)

        self.btn_next = Button(button_frame, text="下一个学生", command=self.next_student, width=15)
        self.btn_next.pack(side=tk.LEFT, padx=5)

        self.chk_mode = Checkbutton(button_frame, text="录像模式", variable=self.mode_var, command=self.toggle_mode)
        self.chk_mode.pack(side=tk.LEFT, padx=5)

        self.rotate_var = IntVar(value=0)
        self.chk_rotate = Checkbutton(button_frame, text="旋转180度", variable=self.rotate_var, command=self.toggle_rotation)
        self.chk_rotate.pack(side=tk.LEFT, padx=5)

        self.volume_scale = Scale(button_frame, from_=0, to=100, orient=tk.HORIZONTAL, label="音量")
        self.volume_scale.set(50)  # 默认音量设置为 50%
        self.volume_scale.pack(side=tk.LEFT, padx=5)

        self.btn_audio_test = Button(button_frame, text="音频测试", command=self.test_audio, width=10)
        self.btn_audio_test.pack(side=tk.LEFT, padx=5)

        self.queue = queue.Queue()
        atexit.register(self.cleanup)

        self.ffmpeg_restart_count = 0
        self.max_ffmpeg_restarts = 3

        self.master.after(100, self.load_excel_data)
        self.master.after(10, self.update)
        self.master.after(100, self.process_queue)

        self.master.bind("<Configure>", self.on_resize)

    def on_resize(self, event):
        self.update_canvas_size()

    def update_canvas_size(self):
        window_width = self.master.winfo_width()
        window_height = self.master.winfo_height()
        canvas_height = window_height - 150
        canvas_width = window_width - 20
        self.canvas.config(width=canvas_width, height=canvas_height)

    def load_excel_data(self):
        print("Starting to load Excel data")
        threading.Thread(target=self._load_excel_data_thread, args=(0,), daemon=True).start()

    def _load_excel_data_thread(self, sheet_index):
        try:
            students_info = load_students_info(self.excel_path, sheet_index)
            self.queue.put(("update_students", students_info))
        except Exception as e:
            print(f"Error in _load_excel_data_thread: {e}")
            traceback.print_exc()
            self.queue.put(("error", str(e)))
        finally:
            self.queue.put(("done", None))

    def process_queue(self):
        try:
            while True:
                message, data = self.queue.get_nowait()
                if message == "update_students":
                    print(f"Updating students: {len(data)} students loaded")
                    self.students_info = data
                    self.update_student_info()
                elif message == "error":
                    print(f"An error occurred: {data}")
                elif message == "done":
                    print("Finished processing queue")
                    break
        except queue.Empty:
            pass
        finally:
            self.master.after(100, self.process_queue)

    def load_sheet(self):
        try:
            sheet_index = int(self.sheet_entry.get())
            print(f"Loading sheet index: {sheet_index}")
            self.label.config(text="正在加载...")
            threading.Thread(target=self._load_excel_data_thread, args=(sheet_index,), daemon=True).start()
        except ValueError:
            print("Invalid sheet index")
            self.label.config(text="无效的Sheet索引")

    def toggle_mode(self):
        print(f"Mode toggled: {'录像' if self.mode_var.get() == 1 else '拍照'}")
        if self.is_recording:
            self.stop_recording()

    def toggle_rotation(self):
        print(f"Rotation toggled: {'开启' if self.rotate_var.get() == 1 else '关闭'}")

    def toggle_recording(self):
        if self.mode_var.get() == 1:  # 录像模式
            if self.is_recording:
                self.stop_recording()
                self.next_student()
            else:
                self.start_recording()
        else:  # 拍照模式
            self.take_snapshot()
            self.next_student()

    def start_recording(self):
        exam_id, name = self.students_info[self.current_student_index]
        self.current_video_name = f"{exam_id}_{name}.mp4"
        
        self.ffmpeg_restart_count = 0
        self._start_ffmpeg_process()

    def _start_ffmpeg_process(self):
        volume = self.volume_scale.get() / 100
        command = [
            "ffmpeg",
            "-f", "avfoundation",
            "-framerate", "30",
            "-video_size", "1280x720",
            "-i", "default",
            "-c:v", "libx264",
            "-preset", "ultrafast",
            "-c:a", "aac",
            "-b:a", "256k",
            "-ar", "48000",
            "-af", f"highpass=f=80,lowpass=f=10000,afftdn=nf=-20,volume={volume}",
            "-movflags", "+faststart",
            "-y",
            self.current_video_name
        ]
        
        if self.rotate_var.get() == 1:
            command.insert(-1, "-vf")
            command.insert(-1, "transpose=2,transpose=2")
        
        try:
            self.ffmpeg_process = subprocess.Popen(command, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            self.is_recording = True
            self.recording_status.config(text="正在录像", fg="red")
            print(f"Recording started: {self.current_video_name}")
        except Exception as e:
            error_message = f"启动录制失败: {str(e)}"
            print(error_message)
            messagebox.showerror("录制失败", error_message)
            self.is_recording = False
            self.recording_status.config(text="就绪", fg="green")

        # 启动一个线程来监控 FFmpeg 进程
        threading.Thread(target=self._monitor_ffmpeg_process, daemon=True).start()

    def _monitor_ffmpeg_process(self):
        while self.is_recording:
            return_code = self.ffmpeg_process.poll()
            if return_code is not None:
                print(f"FFmpeg process ended unexpectedly with return code: {return_code}")
                if self.ffmpeg_restart_count < self.max_ffmpeg_restarts:
                    print("Attempting to restart FFmpeg process...")
                    self.ffmpeg_restart_count += 1
                    self._start_ffmpeg_process()
                else:
                    print("Max restart attempts reached. Stopping recording.")
                    self.stop_recording()
                break
            time.sleep(1)

    def stop_recording(self):
        if self.is_recording and self.ffmpeg_process:
            self.is_recording = False
            try:
                self.ffmpeg_process.stdin.write(b'q')
                self.ffmpeg_process.stdin.flush()
                self.ffmpeg_process.wait(timeout=5)
            except:
                self.ffmpeg_process.terminate()
                self.ffmpeg_process.wait()
            self.ffmpeg_process = None
            print(f"Recording stopped and saved: {self.current_video_name}")
            self.recording_status.config(text="就绪", fg="green")

    def take_snapshot(self):
        ret, frame = self.vid.read()
        if ret:
            if self.rotate_var.get() == 1:
                frame = cv2.rotate(frame, cv2.ROTATE_180)
            exam_id, name = self.students_info[self.current_student_index]
            cv2.imwrite(f"{exam_id}_{name}.png", frame)
            print(f"Photo saved as {exam_id}_{name}.png")

    def next_student(self):
        if self.current_student_index < len(self.students_info) - 1:
            if self.is_recording:
                self.stop_recording()
            self.current_student_index += 1
            self.update_student_info()

    def previous_student(self):
        if self.current_student_index > 0:
            if self.is_recording:
                self.stop_recording()
            self.current_student_index -= 1
            self.update_student_info()

    def update_student_info(self):
        if self.students_info:
            exam_id, name = self.students_info[self.current_student_index]
            self.label.config(text=f"当前学生：{name} ({exam_id})")
        else:
            self.label.config(text="没有学生信息")

    def update(self):
        ret, frame = self.vid.read()
        if ret:
            if self.rotate_var.get() == 1:
                frame = cv2.rotate(frame, cv2.ROTATE_180)
            
            canvas_width = self.canvas.winfo_width()
            canvas_height = self.canvas.winfo_height()

            frame_height, frame_width = frame.shape[:2]
            aspect_ratio = frame_width / frame_height
            
            if canvas_width / canvas_height > aspect_ratio:
                new_height = canvas_height
                new_width = int(new_height * aspect_ratio)
            else:
                new_width = canvas_width
                new_height = int(new_width / aspect_ratio)

            frame_resized = cv2.resize(frame, (new_width, new_height))
            self.photo = ImageTk.PhotoImage(image=Image.fromarray(cv2.cvtColor(frame_resized, cv2.COLOR_BGR2RGB)))
            
            self.canvas.delete("all")
            self.canvas.create_image(canvas_width//2, canvas_height//2, image=self.photo, anchor=tk.CENTER)

        self.master.after(10, self.update)

    def test_audio(self):
        test_command = [
            "ffmpeg",
            "-f", "avfoundation",
            "-framerate", "30",
            "-i", "default",
            "-t", "5",  # 录制5秒
            "-c:a", "aac",
            "-b:a", "256k",
            "-af", "highpass=f=80,lowpass=f=10000,afftdn=nf=-20",
            "-y",
            "audio_test.aac"
        ]
        try:
            result = subprocess.run(test_command, check=True, capture_output=True, text=True)
            print("音频测试完成，请检查 audio_test.aac 文件")
        except subprocess.CalledProcessError as e:
            error_message = f"音频测试失败: {e}\n\n错误输出:\n{e.stderr}"
            print(error_message)
            messagebox.showerror("音频测试失败", error_message)

        # 检查文件是否成功创建
        if os.path.exists("audio_test.aac") and os.path.getsize("audio_test.aac") > 0:
            messagebox.showinfo("音频测试", "音频测试文件已成功创建。请检查 audio_test.aac 文件。")
        else:
            messagebox.showerror("音频测试失败", "无法创建音频测试文件。请检查您的麦克风设置和权限。")

    def cleanup(self):
        print("Cleaning up resources...")
        if self.is_recording:
            self.stop_recording()
        if self.vid.isOpened():
            self.vid.release()
        cv2.destroyAllWindows()
        print("Cleanup completed.")

    def on_closing(self):
        print("Window is closing. Cleaning up...")
        self.cleanup()
        self.master.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    excel_path = "mt.xlsx"
    app = CameraApp(root, excel_path)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    
    try:
        root.mainloop()
    except Exception as e:
        print(f"An error occurred in main loop: {e}")
        traceback.print_exc()
    finally:
        print("Program is exiting. Performing final cleanup...")
        app.cleanup()

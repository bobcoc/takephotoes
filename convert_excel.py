#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文件格式转换程序
将2025.xlsx转换为mt2025.xlsx，使其符合tvds.py的要求
"""

import openpyxl
from openpyxl import Workbook
import os

def convert_excel_format(input_file="2025.xlsx", output_file="mt2025.xlsx"):
    """
    将原始Excel文件转换为tvds.py要求的格式
    
    原始格式：序号、班级、录取编号、考号、新生姓名、性别、备注
    目标格式：考号、姓名（按班级分Sheet）
    """
    
    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        print(f"错误：找不到输入文件 {input_file}")
        return False
    
    try:
        # 打开原始文件
        print(f"正在读取文件：{input_file}")
        workbook = openpyxl.load_workbook(input_file, read_only=True)
        
        # 假设数据在第一个工作表中
        sheet = workbook.active
        
        # 读取所有数据，按班级分组
        class_data = {}
        row_count = 0
        
        print("正在解析数据...")
        
        # 遍历所有行（从第2行开始，跳过表头）
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 5:  # 确保行数据完整
                continue
                
            # 提取数据：序号、班级、录取编号、考号、新生姓名、性别、备注
            serial_num = row[0]  # 序号
            class_num = row[1]   # 班级
            admission_num = row[2]  # 录取编号
            exam_id = row[3]     # 考号
            student_name = row[4]  # 新生姓名
            
            # 检查关键数据是否存在
            if not exam_id or not student_name:
                continue
            
            # 跳过表头行（如果考号列是"考号"文字）
            if str(exam_id).strip() == "考号" or str(student_name).strip() == "新生姓名":
                continue
                
            # 确保班级号是整数
            if class_num is not None:
                try:
                    # 如果班级列包含"班级"文字，跳过
                    if str(class_num).strip() == "班级":
                        continue
                    class_key = f"班级{int(class_num)}"
                except (ValueError, TypeError):
                    # 如果无法转换为整数，使用原值
                    if str(class_num).strip() == "班级":
                        continue
                    class_key = f"班级{class_num}"
            else:
                class_key = "未分班"
            
            # 按班级分组存储
            if class_key not in class_data:
                class_data[class_key] = []
            
            class_data[class_key].append((exam_id, student_name))
            row_count += 1
        
        workbook.close()
        
        print(f"共读取到 {row_count} 条学生记录")
        print(f"发现 {len(class_data)} 个班级：{list(class_data.keys())}")
        
        # 创建新的Excel文件
        print(f"正在创建输出文件：{output_file}")
        new_workbook = Workbook()
        
        # 删除默认的sheet
        new_workbook.remove(new_workbook.active)
        
        # 为每个班级创建一个工作表
        for class_name, students in class_data.items():
            print(f"正在创建工作表：{class_name}（{len(students)}名学生）")
            
            # 创建新的工作表
            ws = new_workbook.create_sheet(title=class_name)
            
            # 添加表头
            ws.append(["考号", "姓名"])
            
            # 添加学生数据
            for exam_id, name in students:
                ws.append([exam_id, name])
            
            # 调整列宽
            ws.column_dimensions['A'].width = 15  # 考号列
            ws.column_dimensions['B'].width = 12  # 姓名列
        
        # 保存文件
        new_workbook.save(output_file)
        new_workbook.close()
        
        print(f"✅ 转换完成！输出文件：{output_file}")
        print("\n文件结构：")
        for class_name, students in class_data.items():
            print(f"  📋 {class_name}: {len(students)}名学生")
        
        return True
        
    except Exception as e:
        print(f"❌ 转换过程中发生错误：{e}")
        import traceback
        traceback.print_exc()
        return False

def verify_output_file(output_file="mt2025.xlsx"):
    """
    验证输出文件的格式是否正确
    """
    print(f"\n正在验证输出文件：{output_file}")
    
    try:
        workbook = openpyxl.load_workbook(output_file, read_only=True)
        sheet_names = workbook.sheetnames
        
        print(f"工作表数量：{len(sheet_names)}")
        
        total_students = 0
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            # 计算学生数量（减去表头行）
            student_count = sheet.max_row - 1 if sheet.max_row > 1 else 0
            total_students += student_count
            print(f"  📋 {sheet_name}: {student_count}名学生")
            
            # 检查前几行数据格式
            if student_count > 0:
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_row=3, values_only=True), 2):
                    if row[0] and row[1]:  # 考号和姓名都不为空
                        print(f"    示例数据: {row[0]} - {row[1]}")
                        break
        
        workbook.close()
        print(f"✅ 验证通过！总计 {total_students} 名学生")
        return True
        
    except Exception as e:
        print(f"❌ 验证失败：{e}")
        return False

if __name__ == "__main__":
    print("🔄 Excel文件格式转换程序")
    print("=" * 50)
    
    # 执行转换
    success = convert_excel_format()
    
    if success:
        # 验证输出文件
        verify_output_file()
        print("\n✨ 转换完成！现在可以在tvds.py中使用mt2025.xlsx文件了。")
    else:
        print("\n❌ 转换失败，请检查输入文件格式。")
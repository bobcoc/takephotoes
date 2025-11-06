#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
学生拍照排查工具
检查名单中还有哪些同学没有拍照
"""

import os
import glob
import openpyxl
import re
from pathlib import Path

def load_all_students_from_excel(excel_path):
    """从Excel文件的所有sheet中加载学生信息"""
    print(f"正在读取Excel文件: {excel_path}")
    workbook = openpyxl.load_workbook(excel_path, read_only=True)
    all_students = []  # 存储所有学生信息 [(考号, 姓名, 班级), ...]
    
    for sheet_name in workbook.sheetnames:
        print(f"正在处理sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2 and row[0] and row[1]:
                exam_id, name = str(row[0]).strip(), str(row[1]).strip()
                if name and exam_id:
                    all_students.append((exam_id, name, sheet_name))
                    print(f"  添加学生: {name} ({exam_id}) - {sheet_name}")
    
    workbook.close()
    print(f"总共加载了 {len(all_students)} 个学生信息")
    return all_students

def get_existing_photos(directory):
    """获取现有的照片文件"""
    pattern = os.path.join(directory, "*.png")
    photo_files = glob.glob(pattern)
    
    existing_photos = set()  # 存储已拍照的学生姓名
    
    for photo_path in photo_files:
        filename = os.path.basename(photo_path)
        # 提取文件名中的学生信息
        name_part = os.path.splitext(filename)[0]
        
        # 匹配模式：考号_姓名
        match = re.match(r'^(\d+)_(.+)$', name_part)
        if match:
            exam_id, name = match.groups()
            existing_photos.add((exam_id, name))
            print(f"找到照片: {name} ({exam_id})")
    
    print(f"总共找到 {len(existing_photos)} 张照片")
    return existing_photos

def check_missing_photos(directory, excel_path):
    """检查缺失照片的主函数"""
    print("="*60)
    print("学生拍照排查工具")
    print("="*60)
    
    # 加载学生名单
    all_students = load_all_students_from_excel(excel_path)
    if not all_students:
        print("❌ 没有从Excel文件中读取到学生信息")
        return
    
    # 获取现有照片
    print(f"\n正在扫描目录: {directory}")
    existing_photos = get_existing_photos(directory)
    
    # 检查缺失的照片
    missing_photos = []
    has_photos = []
    
    for exam_id, name, class_name in all_students:
        if (exam_id, name) in existing_photos:
            has_photos.append((exam_id, name, class_name))
        else:
            missing_photos.append((exam_id, name, class_name))
    
    # 按班级分组显示结果
    print("\n" + "="*60)
    print("拍照情况统计")
    print("="*60)
    
    # 统计各班级情况
    class_stats = {}
    for exam_id, name, class_name in all_students:
        if class_name not in class_stats:
            class_stats[class_name] = {'total': 0, 'has_photo': 0, 'missing': 0}
        class_stats[class_name]['total'] += 1
        
        if (exam_id, name) in existing_photos:
            class_stats[class_name]['has_photo'] += 1
        else:
            class_stats[class_name]['missing'] += 1
    
    # 显示各班级统计
    print("\n各班级拍照统计:")
    print("-" * 60)
    print(f"{'班级':<15} {'总人数':<8} {'已拍照':<8} {'未拍照':<8} {'完成率':<10}")
    print("-" * 60)
    
    for class_name in sorted(class_stats.keys()):
        stats = class_stats[class_name]
        completion_rate = (stats['has_photo'] / stats['total']) * 100 if stats['total'] > 0 else 0
        print(f"{class_name:<15} {stats['total']:<8} {stats['has_photo']:<8} {stats['missing']:<8} {completion_rate:>7.1f}%")
    
    # 显示缺失照片的学生
    if missing_photos:
        print(f"\n❌ 以下 {len(missing_photos)} 位同学还没有拍照:")
        print("-" * 60)
        print(f"{'考号':<12} {'姓名':<10} {'班级':<15}")
        print("-" * 60)
        
        # 按班级分组显示
        missing_by_class = {}
        for exam_id, name, class_name in missing_photos:
            if class_name not in missing_by_class:
                missing_by_class[class_name] = []
            missing_by_class[class_name].append((exam_id, name))
        
        for class_name in sorted(missing_by_class.keys()):
            print(f"\n【{class_name}】:")
            for exam_id, name in sorted(missing_by_class[class_name]):
                print(f"  {exam_id:<12} {name:<10}")
    else:
        print(f"\n✅ 太棒了！所有同学都已经拍照了！")
    
    # 总体统计
    print("\n" + "="*60)
    print("总体统计:")
    print(f"  总学生数: {len(all_students)}")
    print(f"  已拍照数: {len(has_photos)}")
    print(f"  未拍照数: {len(missing_photos)}")
    print(f"  完成率: {(len(has_photos) / len(all_students)) * 100:.1f}%")
    
    # 保存缺失名单到文件
    if missing_photos:
        output_file = os.path.join(directory, "未拍照学生名单.txt")
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write("未拍照学生名单\n")
            f.write("="*40 + "\n\n")
            
            for class_name in sorted(missing_by_class.keys()):
                f.write(f"【{class_name}】\n")
                for exam_id, name in sorted(missing_by_class[class_name]):
                    f.write(f"  {exam_id} {name}\n")
                f.write("\n")
            
            f.write(f"\n统计信息:\n")
            f.write(f"总学生数: {len(all_students)}\n")
            f.write(f"已拍照数: {len(has_photos)}\n")
            f.write(f"未拍照数: {len(missing_photos)}\n")
            f.write(f"完成率: {(len(has_photos) / len(all_students)) * 100:.1f}%\n")
        
        print(f"\n📝 未拍照学生名单已保存到: {output_file}")

def main():
    """主函数"""
    # 设置路径
    current_dir = os.getcwd()
    excel_path = os.path.join(current_dir, "mt2025.xlsx")
    
    # 检查Excel文件是否存在
    if not os.path.exists(excel_path):
        print(f"❌ Excel文件不存在: {excel_path}")
        return
    
    print(f"工作目录: {current_dir}")
    print(f"Excel文件: {excel_path}")
    
    # 执行检查
    check_missing_photos(current_dir, excel_path)

if __name__ == "__main__":
    main()
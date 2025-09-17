#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试转换后的Excel文件是否与tvds.py兼容
"""

import openpyxl

def test_excel_compatibility(excel_path="mt2025.xlsx"):
    """
    测试Excel文件是否与tvds.py兼容
    """
    print(f"🧪 测试文件兼容性：{excel_path}")
    print("=" * 50)
    
    try:
        # 使用与tvds.py相同的方式加载Excel文件
        def load_students_info(excel_path, sheet_index=0):
            print(f"Loading students info from sheet index: {sheet_index}")
            workbook = openpyxl.load_workbook(excel_path, read_only=True)
            sheets = workbook.sheetnames
            if 0 <= sheet_index < len(sheets):
                sheet = workbook[sheets[sheet_index]]
            else:
                print(f"Invalid sheet index: {sheet_index}. Using the first sheet.")
                sheet = workbook.active
            students_info = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                exam_id, name = row[0], row[1]
                if exam_id and name:
                    students_info.append((exam_id, name))
            print(f"Loaded {len(students_info)} students")
            workbook.close()
            return students_info

        def get_sheet_names(excel_path):
            """获取Excel文件中所有sheet的名称"""
            workbook = openpyxl.load_workbook(excel_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
        
        # 获取所有sheet名称
        sheet_names = get_sheet_names(excel_path)
        print(f"📋 工作表列表：{sheet_names}")
        
        # 测试每个sheet
        total_students = 0
        for i, sheet_name in enumerate(sheet_names):
            print(f"\n🔍 测试工作表 {i}: {sheet_name}")
            students = load_students_info(excel_path, i)
            total_students += len(students)
            
            # 显示前几个学生的信息
            if students:
                print(f"  示例学生数据：")
                for j, (exam_id, name) in enumerate(students[:3]):
                    print(f"    {j+1}. 考号：{exam_id}，姓名：{name}")
                if len(students) > 3:
                    print(f"    ... 还有 {len(students) - 3} 名学生")
        
        print(f"\n✅ 兼容性测试通过！")
        print(f"📊 总计：{len(sheet_names)} 个班级，{total_students} 名学生")
        print(f"🎯 该文件可以直接在tvds.py中使用！")
        
        return True
        
    except Exception as e:
        print(f"❌ 兼容性测试失败：{e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    test_excel_compatibility()